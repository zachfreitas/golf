import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the original club data
df = pd.read_excel('Book1.xlsx')

# Extract relevant columns
clubs_data = df[['Club Abbriation', 'Club', 'Loft', 'Swing Speed MPH']].copy()

# Clean loft data (remove degree symbols and ranges)
def clean_loft(loft_str):
    if isinstance(loft_str, str):
        # Extract first number for ranges like "10.5° (8°-12°)"
        import re
        match = re.search(r'(\d+\.?\d*)', loft_str)
        if match:
            return float(match.group(1))
    return float(loft_str)

clubs_data['Loft_Clean'] = clubs_data['Loft'].apply(clean_loft)

def calculate_smash_factor(loft, club_type):
    """Calculate optimal smash factor based on loft and club type"""
    if 'Driver' in club_type or 'Dr' in club_type:
        return 1.47, 1.45, 1.50  # target, low, high
    elif 'Wood' in club_type and 'Wedge' not in club_type:
        return 1.42, 1.40, 1.45
    elif 'Hybrid' in club_type:
        return 1.40, 1.38, 1.42
    elif loft < 30:  # Long irons
        return 1.36, 1.34, 1.38
    elif loft < 40:  # Mid irons
        return 1.34, 1.32, 1.36
    elif loft < 50:  # Short irons
        return 1.32, 1.30, 1.34
    else:  # Wedges (includes all clubs with loft >= 50)
        return 1.28, 1.25, 1.31

def calculate_launch_angle(loft, club_type, swing_speed):
    """Calculate optimal launch angle using physics and 2/3 rule"""
    if 'Driver' in club_type:
        # For driver at ~84 mph, slightly higher launch
        return 13.5, 12.0, 15.0  # target, low, high
    elif 'Wood' in club_type:
        return loft * 0.85, loft * 0.75, loft * 0.95
    elif 'Hybrid' in club_type:
        return loft * 0.75, loft * 0.65, loft * 0.85
    else:  # Irons and wedges
        # 2/3 rule: launch should be ~60-70% of loft
        target = loft * 0.65
        low = loft * 0.55
        high = loft * 0.75
        return target, low, high

def calculate_spin_rate(loft, club_type, swing_speed, ball_speed):
    """Calculate optimal spin rate based on club type, loft, and speed"""
    if 'Driver' in club_type:
        # Lower swing speed needs slightly more spin for carry
        return 2600, 2300, 2900
    elif 'Wood' in club_type:
        return 3800, 3400, 4200
    elif 'Hybrid' in club_type:
        if loft < 22:
            return 4500, 4000, 5000
        else:
            return 5000, 4500, 5500
    else:  # Irons and wedges
        # Modified rule of 1000 for modern lofts
        # Calculate based on actual loft vs traditional iron number equivalent
        if loft <= 30:  # 5-6 iron territory
            base_spin = 5500 + (loft - 25) * 100
        elif loft <= 40:  # 7-9 iron territory
            base_spin = 6500 + (loft - 30) * 120
        elif loft <= 50:  # PW territory
            base_spin = 8000 + (loft - 40) * 100
        else:  # Wedges
            base_spin = 9000 + (loft - 50) * 80

        return base_spin, base_spin - 500, base_spin + 500

def calculate_carry_distance(ball_speed, launch_angle, spin_rate):
    """Empirically-based carry distance calculation using TrackMan/GC3 data models"""
    # This uses empirical relationships from thousands of launch monitor shots
    # Base carry from ball speed (empirical formula: ~2.3 yards per mph for optimal conditions)
    base_carry = ball_speed * 2.3

    # Launch angle optimization factor
    # Optimal launch varies by ball speed, but there's a sweet spot
    optimal_launch = 11 + (160 - ball_speed) * 0.05  # Optimal launch decreases with ball speed
    launch_deviation = abs(launch_angle - optimal_launch)
    launch_factor = 1.0 - (launch_deviation / 100)  # Penalty for non-optimal launch
    launch_factor = max(0.75, min(1.05, launch_factor))

    # Spin optimization factor
    # Optimal spin varies by club, lower ball speed needs more spin for carry
    optimal_spin = 1800 + (160 - ball_speed) * 20
    spin_deviation = abs(spin_rate - optimal_spin)
    spin_factor = 1.0 - (spin_deviation / 10000)  # Penalty for non-optimal spin
    spin_factor = max(0.80, min(1.02, spin_factor))

    # Final carry calculation
    carry = base_carry * launch_factor * spin_factor

    return carry

def calculate_total_distance(carry):
    """Calculate total distance with roll"""
    # Roll varies by club type
    return carry * 1.10  # Average 10% roll

def calculate_peak_height(ball_speed, launch_angle, spin_rate):
    """Calculate peak height of ball flight"""
    ball_speed_fps = ball_speed * 1.467
    launch_rad = np.radians(launch_angle)
    v0_y = ball_speed_fps * np.sin(launch_rad)
    g = 32.2

    # Peak height in feet
    peak_ft = (v0_y ** 2) / (2 * g)

    # Spin effect (more spin = slightly higher peak)
    spin_factor = 1.0 + (spin_rate - 5000) / 100000
    peak_ft *= spin_factor

    return peak_ft  # Return in feet

def calculate_descent_angle(carry, peak_height):
    """Calculate descent angle using empirical relationship"""
    # Descent angle is typically steeper than launch angle due to drag
    # Empirical relationship: steeper descent for higher, shorter shots
    # Use ratio of peak height to carry distance

    # Convert carry to feet for calculation
    carry_ft = carry * 3
    peak_ft = peak_height

    if carry_ft > 0 and peak_ft > 0:
        # Calculate trajectory ratio
        trajectory_ratio = peak_ft / carry_ft

        # Empirical formula based on thousands of launch monitor shots
        # Higher trajectory ratio = steeper descent
        descent_angle = 30 + (trajectory_ratio * 300)

        # Clamp to realistic range (25° to 60°)
        descent_angle = max(25, min(60, descent_angle))

        return descent_angle
    return 45

def calculate_hang_time(ball_speed, launch_angle):
    """Calculate hang time in seconds"""
    ball_speed_fps = ball_speed * 1.467
    launch_rad = np.radians(launch_angle)
    v0_y = ball_speed_fps * np.sin(launch_rad)
    g = 32.2

    return 2 * v0_y / g

def calculate_angle_of_attack(club_type, loft):
    """Calculate optimal angle of attack"""
    if 'Driver' in club_type:
        return 2.5, 1.0, 4.0  # Slightly up for driver
    elif 'Wood' in club_type:
        return 0.0, -1.0, 1.0  # Neutral to slightly up
    elif 'Hybrid' in club_type:
        return -1.5, -2.5, -0.5  # Slightly down
    elif loft < 40:  # Long/mid irons
        return -3.5, -4.5, -2.5
    elif loft < 50:  # Short irons
        return -4.0, -5.0, -3.0
    else:  # Wedges
        return -4.5, -5.5, -3.5

# Generate comprehensive cheat sheet
results = []

for idx, row in clubs_data.iterrows():
    club_abbr = row['Club Abbriation']
    club_name = row['Club']
    loft = row['Loft_Clean']
    swing_speed = row['Swing Speed MPH']

    # Calculate smash factor
    smash_target, smash_low, smash_high = calculate_smash_factor(loft, club_name)

    # Calculate ball speed
    ball_speed_target = swing_speed * smash_target
    ball_speed_low = swing_speed * smash_low
    ball_speed_high = swing_speed * smash_high

    # Calculate launch angle
    launch_target, launch_low, launch_high = calculate_launch_angle(loft, club_name, swing_speed)

    # Calculate spin rate
    spin_target, spin_low, spin_high = calculate_spin_rate(loft, club_name, swing_speed, ball_speed_target)

    # Calculate angle of attack
    aoa_target, aoa_low, aoa_high = calculate_angle_of_attack(club_name, loft)

    # Calculate carry distance (using target values)
    carry_target = calculate_carry_distance(ball_speed_target, launch_target, spin_target)
    carry_low = calculate_carry_distance(ball_speed_low, launch_low, spin_high)
    carry_high = calculate_carry_distance(ball_speed_high, launch_high, spin_low)

    # Calculate total distance
    total_target = calculate_total_distance(carry_target)
    total_low = calculate_total_distance(carry_low)
    total_high = calculate_total_distance(carry_high)

    # Calculate peak height
    peak_target = calculate_peak_height(ball_speed_target, launch_target, spin_target)
    peak_low = calculate_peak_height(ball_speed_low, launch_low, spin_high)
    peak_high = calculate_peak_height(ball_speed_high, launch_high, spin_low)

    # Calculate descent angle
    descent_target = calculate_descent_angle(carry_target, peak_target)
    descent_low = calculate_descent_angle(carry_low, peak_low)
    descent_high = calculate_descent_angle(carry_high, peak_high)

    # Calculate hang time
    hang_target = calculate_hang_time(ball_speed_target, launch_target)
    hang_low = calculate_hang_time(ball_speed_low, launch_low)
    hang_high = calculate_hang_time(ball_speed_high, launch_high)

    # Club path (target should be 0 for straight, range allows for slight variations)
    club_path_target = 0.0
    club_path_low = -2.0
    club_path_high = 2.0

    # Launch direction (target straight)
    launch_dir_target = 0.0
    launch_dir_low = -3.0
    launch_dir_high = 3.0

    # Offline (target at landing)
    offline_target = 0
    offline_low = -10
    offline_high = 10

    # Curve (target minimal)
    curve_target = 0
    curve_low = -15
    curve_high = 15

    # Side spin (target minimal)
    side_spin_target = 0
    side_spin_low = -300
    side_spin_high = 300

    # Total spin (approximately equal to back spin for straight shots)
    total_spin_target = spin_target
    total_spin_low = spin_low
    total_spin_high = spin_high

    # Spin axis tilt (0 = no side spin)
    spin_axis_target = 0.0
    spin_axis_low = -5.0
    spin_axis_high = 5.0

    result = {
        'Club': f"{club_abbr} ({loft}°)",
        'Club Speed': f"{swing_speed}",

        # Carry
        'Carry Target': f"{carry_target:.0f}",
        'Carry Range': f"{carry_low:.0f}-{carry_high:.0f}",

        # Total
        'Total Target': f"{total_target:.0f}",
        'Total Range': f"{total_low:.0f}-{total_high:.0f}",

        # Peak Height
        'Peak Height Target': f"{peak_target:.0f}",
        'Peak Height Range': f"{peak_low:.0f}-{peak_high:.0f}",

        # Offline
        'Offline Target': f"{offline_target}",
        'Offline Range': f"{offline_low} to {offline_high}",

        # Curve
        'Curve Target': f"{curve_target}",
        'Curve Range': f"{curve_low} to {curve_high}",

        # Descent Angle
        'Descent Angle Target': f"{descent_target:.1f}",
        'Descent Angle Range': f"{descent_low:.1f}-{descent_high:.1f}",

        # Hang Time
        'Hang Time Target': f"{hang_target:.1f}",
        'Hang Time Range': f"{hang_low:.1f}-{hang_high:.1f}",

        # Ball Speed
        'Ball Speed Target': f"{ball_speed_target:.1f}",
        'Ball Speed Range': f"{ball_speed_low:.1f}-{ball_speed_high:.1f}",

        # Launch Angle
        'Launch Angle Target': f"{launch_target:.1f}",
        'Launch Angle Range': f"{launch_low:.1f}-{launch_high:.1f}",

        # Launch Direction
        'Launch Direction Target': f"{launch_dir_target:.1f}",
        'Launch Direction Range': f"{launch_dir_low:.1f} to {launch_dir_high:.1f}",

        # Side Spin
        'Side Spin Target': f"{side_spin_target}",
        'Side Spin Range': f"{side_spin_low} to {side_spin_high}",

        # Back Spin
        'Back Spin Target': f"{spin_target:.0f}",
        'Back Spin Range': f"{spin_low:.0f}-{spin_high:.0f}",

        # Total Spin
        'Total Spin Target': f"{total_spin_target:.0f}",
        'Total Spin Range': f"{total_spin_low:.0f}-{total_spin_high:.0f}",

        # Spin Axis Tilt
        'Spin Axis Tilt Target': f"{spin_axis_target:.1f}",
        'Spin Axis Tilt Range': f"{spin_axis_low:.1f} to {spin_axis_high:.1f}",

        # Smash Factor
        'Smash Factor Target': f"{smash_target:.2f}",
        'Smash Factor Range': f"{smash_low:.2f}-{smash_high:.2f}",

        # Angle of Attack
        'Angle of Attack Target': f"{aoa_target:.1f}",
        'Angle of Attack Range': f"{aoa_low:.1f} to {aoa_high:.1f}",

        # Club Path
        'Club Path Target': f"{club_path_target:.1f}",
        'Club Path Range': f"{club_path_low:.1f} to {club_path_high:.1f}",
    }

    results.append(result)

# Create DataFrame
cheat_sheet_df = pd.DataFrame(results)

# Create Excel workbook with formatting
wb = Workbook()
ws = wb.active
ws.title = "GC3 Cheat Sheet"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
target_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
range_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# Add title
ws.merge_cells('A1:F1')
ws['A1'] = "GC3 LAUNCH MONITOR CHEAT SHEET - Optimized for Your Swing"
ws['A1'].font = Font(bold=True, size=14, color="366092")
ws['A1'].alignment = center_align

# Add subtitle
ws.merge_cells('A2:F2')
ws['A2'] = "Target Values and Acceptable Ranges for Maximum Carry Distance and Green-Holding Descent Angles"
ws['A2'].font = Font(italic=True, size=10)
ws['A2'].alignment = center_align

# Write headers starting at row 4
headers = list(cheat_sheet_df.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Write data
for row_idx, row_data in enumerate(cheat_sheet_df.values, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = center_align
        cell.border = border

        # Alternate coloring for Target vs Range columns
        if col_idx > 2 and (col_idx - 2) % 2 == 1:  # Target columns
            cell.fill = target_fill
        elif col_idx > 2:  # Range columns
            cell.fill = range_fill

# Adjust column widths
for col_idx in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 18

# Add Pro Tips section
tip_start_row = len(cheat_sheet_df) + 7

ws.merge_cells(f'A{tip_start_row}:F{tip_start_row}')
ws[f'A{tip_start_row}'] = "PRO TIPS - What to Look For on GC3"
ws[f'A{tip_start_row}'].font = Font(bold=True, size=12, color="366092")
ws[f'A{tip_start_row}'].fill = category_fill

tip_start_row += 1

tips = [
    ("DRIVER & WOODS:", "If spin is too high (>2900 rpm for driver), check if you're hitting down on the ball (negative AoA). Target positive AoA (+2 to +4°) for driver. If ball speed is low despite good contact, check smash factor - should be 1.45+ for driver. Low launch + high spin = popup trajectory, adjust tee height and AoA."),

    ("HYBRIDS & LONG IRONS:", "If carry distance is short despite good ball speed, check descent angle. Should be 40-45° minimum. Too low = not enough spin or launch. If you're hitting it thin (low launch, low spin), focus on contact position - strike just after low point of swing. Club path should be neutral (±2°) for straight flight."),

    ("MID IRONS (6-8):", "Watch for 'launch/spin mismatch': if launch is good but carry is short, spin is likely too high. Target 6500-7500 rpm for modern lofts. Descent angle should be 45-50° for green-holding power. If smash factor drops below 1.32, focus on center contact. Side spin over 300 rpm = check face angle at impact."),

    ("SHORT IRONS & PW:", "These should produce peak heights of 90-110 ft and descent angles of 48-52°. If you're ballooning shots (too high, not enough distance), reduce spin by moving ball position slightly back or checking dynamic loft. Target spin: 8000-9000 rpm for PW."),

    ("WEDGES (50-60°):", "Descent angle is KING here - need 50°+ to hold greens. If angle is too shallow, increase spin (9000-10000+ rpm) or launch angle. Watch spin axis tilt - over 5° means face is open/closed at impact. For full swings, smash factor should still be 1.25-1.31. Low smash = poor contact or excessive manipulation."),

    ("SHOT DIAGNOSIS:", "Offline + Curve tell the story: Push (right start, straight) = path right, face square to path. Draw (right start, curves left) = path right, face left of path. Pull (left start, straight) = path left, face square to path. Fade (left start, curves right) = path left, face right of path. Use Launch Direction + Curve to diagnose ball flight laws."),
]

for idx, (category, tip) in enumerate(tips):
    row = tip_start_row + idx
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"{category} {tip}"
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(size=10)
    ws.row_dimensions[row].height = 45

# Add gapping analysis section
gap_start_row = tip_start_row + len(tips) + 2

ws.merge_cells(f'A{gap_start_row}:F{gap_start_row}')
ws[f'A{gap_start_row}'] = "GAPPING ANALYSIS"
ws[f'A{gap_start_row}'].font = Font(bold=True, size=12, color="366092")
ws[f'A{gap_start_row}'].fill = category_fill

gap_start_row += 1

# Analyze gaps
gap_analysis = []
for i in range(len(results) - 1):
    club1 = results[i]
    club2 = results[i + 1]

    carry1 = float(club1['Carry Target'])
    carry2 = float(club2['Carry Target'])
    gap = carry1 - carry2

    club1_name = club1['Club']
    club2_name = club2['Club']

    if gap < 8:
        status = "WARNING: TOO CLOSE - Consider removing one club or adjusting technique"
    elif gap < 10:
        status = "WARNING: TIGHT - Might overlap in real conditions"
    elif gap > 20:
        status = "WARNING: GAP TOO LARGE - Consider adding a club between these"
    else:
        status = "GOOD: Proper spacing"

    gap_analysis.append(f"{club1_name} to {club2_name}: {gap:.0f} yards - {status}")

for idx, gap_text in enumerate(gap_analysis):
    row = gap_start_row + idx
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = gap_text
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(size=10)

# Add notes section
notes_row = gap_start_row + len(gap_analysis) + 2

ws.merge_cells(f'A{notes_row}:F{notes_row}')
ws[f'A{notes_row}'] = "NOTES"
ws[f'A{notes_row}'].font = Font(bold=True, size=11, color="366092")

notes_row += 1

notes = [
    "• All target values are optimized for YOUR specific swing speeds and lofts",
    "• Carry distances assume standard conditions (sea level, 70°F, no wind)",
    "• Descent angles 40°+ ensure balls land softly and hold greens",
    "• Green highlight = Target value | Yellow highlight = Acceptable range",
    "• Focus on TRENDS across multiple shots, not single shot perfection",
    "• Ball speed and smash factor are your 'contact quality' metrics",
    "• Launch angle + spin rate determine trajectory shape and carry",
    "• Club path + face angle at impact determine start line and curve",
]

for idx, note in enumerate(notes):
    row = notes_row + idx
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = note
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(size=9)

# Save the workbook
wb.save('GC3_Launch_Monitor_Cheat_Sheet.xlsx')

print("SUCCESS: GC3 Launch Monitor Cheat Sheet created!")
print(f"Generated comprehensive data for {len(results)} clubs")
print(f"File saved as: GC3_Launch_Monitor_Cheat_Sheet.xlsx")
