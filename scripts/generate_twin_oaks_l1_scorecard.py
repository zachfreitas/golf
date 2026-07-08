"""Generate a printable Level 1 Scoring Method scorecard for Twin Oaks GC front nine.

White tees | Level 1 = 100-yard scoring zone
Output: outputs/twin_oaks_front9_L1_scorecard.xlsx
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = REPO_ROOT / "outputs" / "twin_oaks_front9_L1_scorecard.xlsx"

# ── Palette ─────────────────────────────────────────────────────────────────
GREEN_DARK   = "1A5276"   # header bg
GREEN_LIGHT  = "D5F5E3"   # scoring-opportunity accent
RED_LIGHT    = "FADBD8"   # nemesis hole accent
BLUE_LIGHT   = "D6EAF8"   # par-3 accent
YELLOW_LIGHT = "FEF9E7"   # par-5 accent
GREY_LIGHT   = "F2F3F4"   # alternating rows
GREY_MID     = "D5D8DC"   # totals row
WHITE        = "FFFFFF"
GOLD         = "F39C12"   # nemesis star
ORANGE       = "F0B27A"   # zone goal cells

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

def border(thick=False):
    s = "medium" if thick else "thin"
    side = Side(style=s)
    return Border(left=side, right=side, top=side, bottom=side)

def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Course data ──────────────────────────────────────────────────────────────
HOLES = [
    # (hole, yards, par, handicap, row_type, tee_club, zone_goal, strategy_tip)
    # row_type: "nemesis", "opportunity", "par3", "par5", "standard"
    (1, 344, 4, 7,  "nemesis",     "Driver / 3W",   "≤2 shots",
     "NEMESIS (avg +1.6) — keep it IN PLAY above all. Consider 3W if driver brings trouble."),
    (2, 475, 5, 3,  "par5",        "Driver",         "≤2 shots",
     "Driver + 4H/3H gets inside zone in 2 (~383-399 yd). 52% blow-up rate — don't force it."),
    (3, 173, 3, 17, "nemesis",     "3-Hybrid",       "1 shot",
     "NEMESIS (52% dbl+) — 173 yd forces a full 3H. Tee shot MUST reach zone (<100 yd). Aim center green, do not short-side."),
    (4, 352, 4, 11, "standard",    "Driver",         "≤2 shots",
     "Driver leaves ~135 yd → 7i/8i to zone. Standard execution. Commit to the approach."),
    (5, 501, 5, 1,  "opportunity", "Driver",         "≤2 shots",
     "SCORING HOLE — Driver + 3W (~415 yd) puts you inside zone in 2 with 86 yd left. Birdie look."),
    (6, 333, 4, 13, "standard",    "Driver",         "≤2 shots",
     "Short par 4. Driver → ~116 yd → GW/52°. Birdie chance. Lay up to full wedge distance if in trouble."),
    (7, 142, 3, 15, "par3",        "5i / 6i",        "1 shot",
     "Short par 3 — 142 yd is 5i/6i range. Land on green in 1, D3 for easy par. Aim center, avoid back bunker."),
    (8, 319, 4, 9,  "standard",    "Driver",         "1–2 shots",
     "Shortest par 4. Driver → ~100 yd → wedge. Could reach zone in 1 with long drive. Scoring chance."),
    (9, 353, 4, 5,  "nemesis",     "Driver / 3W",    "≤2 shots",
     "NEMESIS (avg +1.5) — finishing hole pressure. Stay disciplined: in play first, distance second."),
]

# ── L1 column headers ────────────────────────────────────────────────────────
# Primary track: In Play | To Zone | In Zone | D3 | Bogey ✓ | Score | Putts | Pen
TRACKING_HEADERS = ["In Play\n(Y/N)", "To\nZone #", "In\nZone #", "D3\n(Y/N)", "Bogey\nCeil ✓", "Score", "Putts", "Pen"]

# ── Bogey ceiling lookup ─────────────────────────────────────────────────────
BOGEY_CEILING = {3: 4, 4: 5, 5: 6}


def build_scorecard():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Front Nine"

    # Page setup for printing
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── Column widths ────────────────────────────────────────────────────────
    # A=Hole, B=Yds, C=Par, D=HCP, E=Tee Club, F=Zone Goal, G..N=tracking, O=Strategy
    col_widths = {
        "A": 6,    # Hole
        "B": 6,    # Yds
        "C": 5,    # Par
        "D": 5,    # HCP
        "E": 12,   # Tee Club
        "F": 9,    # Zone Goal
        "G": 8,    # In Play
        "H": 7,    # To Zone
        "I": 7,    # In Zone
        "J": 7,    # D3
        "K": 8,    # Bogey Ceil
        "L": 7,    # Score
        "M": 7,    # Putts
        "N": 5,    # Pen
        "O": 48,   # Strategy
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Row heights ──────────────────────────────────────────────────────────
    # Will be set as we build

    row = 1

    # ── Banner ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:N{row}")
    cell = ws[f"A{row}"]
    cell.value = "TWIN OAKS GOLF COURSE — FRONT NINE   |   WHITE TEES   |   Scoring Method Level 1"
    cell.font = font(bold=True, size=13, color=WHITE)
    cell.fill = fill(GREEN_DARK)
    cell.alignment = center()
    ws.row_dimensions[row].height = 22
    # Strategy column header (same banner)
    ws[f"O{row}"].fill = fill(GREEN_DARK)
    row += 1

    # ── Sub-header: rating + L1 promise ─────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "White: 70.0 / 126   |   Par 36   |   2,992 yds"
    cell.font = font(size=9, italic=True)
    cell.fill = fill("2E86C1")
    cell.font = Font(size=9, italic=True, color=WHITE)
    cell.alignment = center()

    ws.merge_cells(f"G{row}:N{row}")
    cell = ws[f"G{row}"]
    cell.value = 'L1 PROMISE: Keep ball IN PLAY + reach 100-yd zone + get DOWN IN 3 = BOGEY CEILING on every hole'
    cell.font = Font(size=8, bold=True, color=WHITE)
    cell.fill = fill("1A5276")
    cell.alignment = center(wrap=False)

    ws[f"O{row}"].value = "Date: ___________   Player: ___________________   HCP: _____"
    ws[f"O{row}"].font = font(size=9, bold=True)
    ws[f"O{row}"].alignment = left(wrap=False)
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Column headers ───────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 32

    header_labels = ["Hole", "Yds", "Par", "HCP", "Tee Club", "Zone\nGoal"] + TRACKING_HEADERS + ["Strategy Notes (Twin Oaks — Level 1)"]
    header_cols   = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"]
    header_fills  = [GREEN_DARK]*6 + ["1A5276"]*8 + [GREEN_DARK]

    for col_letter, label, hfill in zip(header_cols, header_labels, header_fills):
        c = ws[f"{col_letter}{row}"]
        c.value = label
        c.font = font(bold=True, size=9, color=WHITE)
        c.fill = fill(hfill)
        c.alignment = center(wrap=True)
        c.border = thin_border()
    row += 1

    # ── Hole rows ─────────────────────────────────────────────────────────────
    total_yds = 0
    for i, (hole, yds, par, hcp, row_type, tee_club, zone_goal, strategy) in enumerate(HOLES):
        ws.row_dimensions[row].height = 40

        # Row background by hole type
        if row_type == "nemesis":
            bg = RED_LIGHT
        elif row_type == "opportunity":
            bg = GREEN_LIGHT
        elif row_type == "par3":
            bg = BLUE_LIGHT
        elif row_type == "par5":
            bg = YELLOW_LIGHT
        else:
            bg = GREY_LIGHT if i % 2 == 0 else WHITE

        bogey_ceil = BOGEY_CEILING[par]
        total_yds += yds

        row_data = [hole, yds, par, hcp, tee_club, zone_goal]
        for ci, (col_letter, val) in enumerate(zip(["A","B","C","D","E","F"], row_data)):
            c = ws[f"{col_letter}{row}"]
            c.value = val
            c.fill = fill(bg)
            c.border = thin_border()
            if col_letter in ("E", "F"):
                c.font = font(size=9, bold=(col_letter == "F"))
                c.alignment = center(wrap=True)
            else:
                c.font = font(size=10, bold=(col_letter in ("A","C")))
                c.alignment = center()

        # Zone Goal cell gets an orange tint for emphasis
        ws[f"F{row}"].fill = fill(ORANGE)

        # Tracking columns (G-N) — blank for user to fill in on course
        for col_letter in ["G","H","I","J","K","L","M","N"]:
            c = ws[f"{col_letter}{row}"]
            c.fill = fill(WHITE)
            c.border = thin_border()
            c.alignment = center()

        # Strategy column
        c = ws[f"O{row}"]
        c.value = strategy
        c.fill = fill(bg)
        c.font = font(size=8.5, italic=(row_type == "nemesis"), bold=(row_type == "nemesis"))
        c.alignment = left(wrap=True)
        c.border = thin_border()

        row += 1

    # ── Totals row ───────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    totals = [("A", "OUT"), ("B", total_yds), ("C", 36), ("D", ""), ("E", ""), ("F", "")]
    for col_letter, val in totals:
        c = ws[f"{col_letter}{row}"]
        c.value = val
        c.font = font(bold=True, size=10)
        c.fill = fill(GREY_MID)
        c.alignment = center()
        c.border = thin_border()
    for col_letter in ["G","H","I","J","K","L","M","N"]:
        c = ws[f"{col_letter}{row}"]
        c.fill = fill(GREY_MID)
        c.border = thin_border()
        c.alignment = center()
    ws[f"O{row}"].fill = fill(GREY_MID)
    row += 1

    # ── L1 Reference Legend ──────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"A{row}:N{row}")
    ws[f"A{row}"].value = (
        "L1 ZONE TARGETS   |   Par 3: ≤1 shot to zone + ≤3 in zone = Bogey Ceiling (4)   "
        "|   Par 4: ≤2 shots to zone + ≤3 in zone = Bogey Ceiling (5)   "
        "|   Par 5: ≤3 shots to zone + ≤3 in zone = Bogey Ceiling (6)   "
        "|   Zone = Inside 100 yd from pin"
    )
    ws[f"A{row}"].font = Font(size=8, bold=True, color=WHITE)
    ws[f"A{row}"].fill = fill("1A5276")
    ws[f"A{row}"].alignment = center(wrap=False)
    ws[f"O{row}"].fill = fill("1A5276")
    row += 1

    # ── Bag reference (distances) ────────────────────────────────────────────
    ws.row_dimensions[row].height = 13
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "YOUR BAG (White Smart Dist):"
    ws[f"A{row}"].font = font(bold=True, size=8)
    ws[f"A{row}"].fill = fill(GREY_LIGHT)
    ws[f"A{row}"].alignment = center()

    bag_cols = [
        ("G", "Dr 217"),
        ("H", "3W 198"),
        ("I", "3H 182"),
        ("J", "4H 166"),
        ("K", "5i 153"),
        ("L", "6i 141"),
        ("M", "7i 129"),
        ("N", "8i 124"),
    ]
    for col_letter, val in bag_cols:
        c = ws[f"{col_letter}{row}"]
        c.value = val
        c.font = font(size=8, bold=True)
        c.fill = fill(GREY_LIGHT)
        c.alignment = center()
        c.border = thin_border()

    ws[f"O{row}"].value = "9i 108 | PW 107 | 52° 75 | 56° 65 | 60° 46"
    ws[f"O{row}"].font = font(size=8, bold=True)
    ws[f"O{row}"].fill = fill(GREY_LIGHT)
    ws[f"O{row}"].alignment = left(wrap=False)
    row += 1

    # ── Color legend ─────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 13
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "Red = Nemesis"
    ws[f"A{row}"].fill = fill(RED_LIGHT)
    ws[f"A{row}"].font = font(size=8, bold=True)
    ws[f"A{row}"].alignment = center()

    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"].value = "Green = Opportunity"
    ws[f"C{row}"].fill = fill(GREEN_LIGHT)
    ws[f"C{row}"].font = font(size=8, bold=True)
    ws[f"C{row}"].alignment = center()

    ws.merge_cells(f"E{row}:F{row}")
    ws[f"E{row}"].value = "Blue = Par 3"
    ws[f"E{row}"].fill = fill(BLUE_LIGHT)
    ws[f"E{row}"].font = font(size=8, bold=True)
    ws[f"E{row}"].alignment = center()

    ws.merge_cells(f"G{row}:H{row}")
    ws[f"G{row}"].value = "Yellow = Par 5"
    ws[f"G{row}"].fill = fill(YELLOW_LIGHT)
    ws[f"G{row}"].font = font(size=8, bold=True)
    ws[f"G{row}"].alignment = center()

    ws.merge_cells(f"I{row}:N{row}")
    ws[f"I{row}"].value = "TRACKING: In Play = no penalty | D3 = strokes in zone ≤3 | Bogey Ceil ✓ = worst score is bogey"
    ws[f"I{row}"].font = font(size=8)
    ws[f"I{row}"].alignment = left(wrap=False)

    ws[f"O{row}"].value = "thescoringmethod.com | Will Robins Level 1"
    ws[f"O{row}"].font = Font(size=8, italic=True, color="707070")
    ws[f"O{row}"].alignment = left(wrap=False)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    build_scorecard()
